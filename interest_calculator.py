"""
利息计算器 - 多轮还款明细表
分段计息，支持LPR浮动利率
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from tkcalendar import DateEntry
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
import os
import urllib.request
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 常量定义
DATE_START_LEGAL = date(2015, 9, 1)
DATE_START_LPR = date(2020, 8, 20)

# 法定利率上限
RATE_CAP_PRE_2015 = Decimal("0.24")
RATE_CAP_2015_TO_2020_FULL = Decimal("0.36")
RATE_CAP_2015_TO_2020_PARTIAL = Decimal("0.24")

# 内置默认LPR值（当网络获取失败时使用）
DEFAULT_LPR = Decimal("3.45")  # 3.45% (percentage format)

# 内置LPR历史数据（2019年8月至今主要记录）
BUILTIN_LPR_DATA = [
    {'date': '2019-08-20', 'lpr': 4.31},
    {'date': '2019-09-20', 'lpr': 4.20},
    {'date': '2019-10-21', 'lpr': 4.20},
    {'date': '2019-11-20', 'lpr': 4.15},
    {'date': '2019-12-20', 'lpr': 4.15},
    {'date': '2020-01-20', 'lpr': 4.15},
    {'date': '2020-02-20', 'lpr': 4.05},
    {'date': '2020-03-20', 'lpr': 4.05},
    {'date': '2020-04-20', 'lpr': 3.85},
    {'date': '2020-05-20', 'lpr': 3.85},
    {'date': '2020-06-22', 'lpr': 3.85},
    {'date': '2020-07-20', 'lpr': 3.85},
    {'date': '2020-08-20', 'lpr': 3.85},
    {'date': '2020-09-21', 'lpr': 3.85},
    {'date': '2020-10-20', 'lpr': 3.85},
    {'date': '2020-11-20', 'lpr': 3.85},
    {'date': '2020-12-21', 'lpr': 3.85},
    {'date': '2021-01-20', 'lpr': 3.85},
    {'date': '2021-02-20', 'lpr': 3.85},
    {'date': '2021-03-22', 'lpr': 3.85},
    {'date': '2021-04-20', 'lpr': 3.85},
    {'date': '2021-05-20', 'lpr': 3.85},
    {'date': '2021-06-21', 'lpr': 3.85},
    {'date': '2021-07-20', 'lpr': 3.85},
    {'date': '2021-08-20', 'lpr': 3.85},
    {'date': '2021-09-22', 'lpr': 3.85},
    {'date': '2021-10-20', 'lpr': 3.85},
    {'date': '2021-11-22', 'lpr': 3.85},
    {'date': '2021-12-20', 'lpr': 3.80},
    {'date': '2022-01-20', 'lpr': 3.70},
    {'date': '2022-02-21', 'lpr': 3.70},
    {'date': '2022-03-21', 'lpr': 3.70},
    {'date': '2022-04-20', 'lpr': 3.70},
    {'date': '2022-05-20', 'lpr': 3.70},
    {'date': '2022-06-20', 'lpr': 3.70},
    {'date': '2022-07-20', 'lpr': 3.70},
    {'date': '2022-08-22', 'lpr': 3.65},
    {'date': '2022-09-20', 'lpr': 3.65},
    {'date': '2022-10-20', 'lpr': 3.65},
    {'date': '2022-11-21', 'lpr': 3.65},
    {'date': '2022-12-20', 'lpr': 3.65},
    {'date': '2023-01-20', 'lpr': 3.65},
    {'date': '2023-02-20', 'lpr': 3.65},
    {'date': '2023-03-20', 'lpr': 3.65},
    {'date': '2023-04-20', 'lpr': 3.65},
    {'date': '2023-05-22', 'lpr': 3.65},
    {'date': '2023-06-20', 'lpr': 3.55},
    {'date': '2023-07-20', 'lpr': 3.55},
    {'date': '2023-08-21', 'lpr': 3.45},
    {'date': '2023-09-20', 'lpr': 3.45},
    {'date': '2023-10-20', 'lpr': 3.45},
    {'date': '2023-11-20', 'lpr': 3.45},
    {'date': '2023-12-20', 'lpr': 3.45},
    {'date': '2024-01-22', 'lpr': 3.45},
    {'date': '2024-02-20', 'lpr': 3.45},
    {'date': '2024-03-20', 'lpr': 3.45},
    {'date': '2024-04-22', 'lpr': 3.45},
    {'date': '2024-05-20', 'lpr': 3.45},
    {'date': '2024-06-20', 'lpr': 3.45},
    {'date': '2024-07-22', 'lpr': 3.35},
    {'date': '2024-08-20', 'lpr': 3.35},
    {'date': '2024-09-20', 'lpr': 3.35},
    {'date': '2024-10-21', 'lpr': 3.35},
    {'date': '2024-11-20', 'lpr': 3.35},
    {'date': '2024-12-20', 'lpr': 3.35},
    {'date': '2025-01-20', 'lpr': 3.35},
    {'date': '2025-02-20', 'lpr': 3.35},
    {'date': '2025-03-20', 'lpr': 3.35},
    {'date': '2025-04-21', 'lpr': 3.35},
    {'date': '2025-05-20', 'lpr': 3.00},
    {'date': '2025-06-20', 'lpr': 3.00},
    {'date': '2025-07-21', 'lpr': 3.00},
    {'date': '2025-08-20', 'lpr': 3.00},
    {'date': '2025-09-22', 'lpr': 3.00},
    {'date': '2025-10-20', 'lpr': 3.00},
    {'date': '2025-11-20', 'lpr': 3.00},
    {'date': '2025-12-22', 'lpr': 3.00},
    {'date': '2026-01-20', 'lpr': 3.00},
    {'date': '2026-02-24', 'lpr': 3.00},
    {'date': '2026-03-20', 'lpr': 3.00},
    {'date': '2026-04-20', 'lpr': 3.00},
    {'date': '2026-05-20', 'lpr': 3.00},
]

# LPR数据文件
LPR_DATA_FILE = "lpr_data.json"
CONFIG_FILE = "config.json"
MONEY_PLACES = Decimal("0.01")
RATE_PLACES = Decimal("0.0001")
YEAR_DAYS = Decimal("360")
LPR_FETCH_ERROR = ""


def to_decimal(value, default=None):
    """将用户输入或数字安全转换为Decimal。"""
    if value is None:
        if default is not None:
            return default
        raise ValueError("empty decimal")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip()
    if not text:
        if default is not None:
            return default
        raise ValueError("empty decimal")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"invalid decimal: {value}") from exc


def money_text(value):
    return format(to_decimal(value, Decimal("0")).quantize(MONEY_PLACES, rounding=ROUND_HALF_UP), "f")


def percent_text(rate):
    percent = (to_decimal(rate, Decimal("0")) * Decimal("100")).quantize(RATE_PLACES, rounding=ROUND_HALF_UP)
    return format(percent, "f").rstrip("0").rstrip(".")


def decimal_to_excel_number(value):
    return float(to_decimal(value, Decimal("0")).quantize(MONEY_PLACES, rounding=ROUND_HALF_UP))


def get_app_data_dir():
    """优先使用用户目录，避免安装目录不可写。"""
    local_app_data = os.environ.get("LOCALAPPDATA")
    if local_app_data:
        cache_dir = os.path.join(local_app_data, "利息计算器")
        try:
            os.makedirs(cache_dir, exist_ok=True)
            return cache_dir
        except OSError:
            pass
    return "."


def get_cache_lpr_path():
    return os.path.join(get_app_data_dir(), LPR_DATA_FILE)


def get_config_path():
    return os.path.join(get_app_data_dir(), CONFIG_FILE)


def load_config():
    path = get_config_path()
    if not os.path.exists(path):
        return {}
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_config(config):
    try:
        with open(get_config_path(), 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"保存配置失败: {e}")
        return False


def get_tushare_token():
    token = os.environ.get("TUSHARE_TOKEN", "").strip()
    if token:
        return token
    return str(load_config().get("tushare_token", "")).strip()


def save_tushare_token(token):
    config = load_config()
    config["tushare_token"] = token.strip()
    return save_config(config)


def days_between(d1, d2):
    """计算两个日期之间的天数"""
    return (d2 - d1).days


def get_rate_for_period(start_date, end_date, user_rate, use_lpr, lpr_multiplier, lpr_data):
    """
    根据时间段获取适用的实际利率
    规则：实际利率 = min(用户输入利率, 法定上限)
    """
    return get_rate_info_for_period(
        start_date, end_date, user_rate, use_lpr, lpr_multiplier, lpr_data
    )["rate"]


def get_rate_info_for_period(start_date, end_date, user_rate, use_lpr, lpr_multiplier, lpr_data):
    """返回指定半开区间[start_date, end_date)的利率和上限信息。"""
    user_rate = to_decimal(user_rate, Decimal("0"))
    lpr_multiplier = to_decimal(lpr_multiplier, Decimal("4"))
    cap = get_legal_cap(start_date, end_date)
    cap_parts = []
    lpr_rate = None
    lpr_cap = None

    if cap is not None:
        cap_parts.append(f"法定上限{percent_text(cap)}%")

    # 2020-08-20后法定上限为LPR倍数；勾选LPR时，早期区间也按更低上限控制。
    if start_date >= DATE_START_LPR or use_lpr:
        lpr_rate = get_lpr_rate_for_date(start_date, lpr_data)
        lpr_cap = lpr_rate * lpr_multiplier
        cap = lpr_cap if cap is None else min(cap, lpr_cap)
        cap_parts.append(f"LPR {percent_text(lpr_rate)}% × {lpr_multiplier} = {percent_text(lpr_cap)}%")

    if cap is None:
        cap = user_rate

    actual_rate = min(user_rate, cap)
    return {
        "rate": actual_rate,
        "cap": cap,
        "cap_description": "；".join(cap_parts) if cap_parts else "无单独上限",
        "lpr_rate": lpr_rate,
        "lpr_cap": lpr_cap,
    }


def get_legal_cap(start_date, end_date):
    """
    根据时间段判断法定利率上限
    判断逻辑基于半开区间[start_date, end_date)
    """
    # 起点在2015-09-01之前，终点任意 -> 24%
    if start_date < DATE_START_LEGAL:
        return RATE_CAP_PRE_2015

    # 起点在2015-09-01~2020-08-19区间
    if DATE_START_LEGAL <= start_date < DATE_START_LPR:
        # 终点未跨过2020-08-20 -> 36%
        if end_date <= DATE_START_LPR:
            return RATE_CAP_2015_TO_2020_FULL
        # 终点超出区间 -> 24%
        else:
            return RATE_CAP_2015_TO_2020_PARTIAL

    # 起点在2020-08-20之后 -> LPR×4（在调用处动态计算）
    if start_date >= DATE_START_LPR:
        return None

    return RATE_CAP_PRE_2015  # 默认


def get_lpr_rate_for_date(target_date, lpr_data):
    """获取指定日期适用的LPR值（找最近的但不晚于该日期的LPR）"""
    if not lpr_data:
        return DEFAULT_LPR / Decimal("100")  # 转换为小数

    # 按日期升序排列，找到最近的不晚于目标日期的LPR
    applicable_lpr = DEFAULT_LPR / Decimal("100")
    for record in sorted(lpr_data, key=lambda item: item.get('date', '')):
        record_date = datetime.strptime(record['date'], '%Y-%m-%d').date()
        if record_date <= target_date:
            applicable_lpr = to_decimal(record['lpr']) / Decimal("100")  # 转换为小数
        else:
            break

    return applicable_lpr


def fetch_chinamoney_lpr_data(start_date=date(2019, 8, 1), end_date=None):
    """
    从中国货币网公开接口获取LPR历史数据。
    该接口一次最多查询一年，因此按时间段拆分请求。
    """
    if end_date is None:
        end_date = date.today()

    records_by_date = {}
    current_start = start_date
    while current_start <= end_date:
        current_end = min(current_start + timedelta(days=360), end_date)
        url = (
            "https://www.chinamoney.com.cn/ags/ms/cm-u-bk-currency/LprHis"
            f"?lang=CN&strStartDate={current_start.strftime('%Y-%m-%d')}"
            f"&strEndDate={current_end.strftime('%Y-%m-%d')}"
        )
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode('utf-8'))

        message = result.get('data', {}).get('message')
        if message:
            raise RuntimeError(f"中国货币网返回: {message}")

        for item in result.get('records', []):
            record_date = item.get('showDateCN')
            lpr_val = item.get('1Y')
            if record_date and lpr_val:
                records_by_date[record_date] = {
                    'date': record_date,
                    'lpr': float(lpr_val)
                }

        current_start = current_end + timedelta(days=1)

    lpr_list = sorted(records_by_date.values(), key=lambda x: x['date'])
    if not lpr_list:
        raise RuntimeError("中国货币网未返回LPR记录")
    return lpr_list


def fetch_tushare_lpr_data(token=None):
    """
    获取LPR历史数据
    使用Tushare Pro API: shibor_lpr
    """
    try:
        token = (token or get_tushare_token()).strip()
        if not token:
            return None
        url = 'https://api.tushare.pro/'
        params = {
            'api_name': 'shibor_lpr',
            'token': token,
            'params': {'start_date': '20190801', 'end_date': '20261231'}
        }

        req = urllib.request.Request(
            url,
            data=json.dumps(params).encode('utf-8'),
            headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'},
            method='POST'
        )

        with urllib.request.urlopen(req, timeout=30) as response:
            result = json.loads(response.read().decode('utf-8'))
            if result.get('code') not in (None, 0):
                raise RuntimeError(result.get('msg') or f"Tushare错误代码: {result.get('code')}")
            items = result.get('data', {}).get('items', [])

            lpr_list = []
            for item in items:
                # Without fields param: ['20190820', 4.25, 4.85]
                date_val = item[0]
                lpr_val = item[1]
                if (isinstance(date_val, str) and len(date_val) == 8 and
                    lpr_val is not None and isinstance(lpr_val, (int, float))):
                    lpr_list.append({
                        'date': f'{date_val[:4]}-{date_val[4:6]}-{date_val[6:8]}',
                        'lpr': float(lpr_val)
                    })

            lpr_list.sort(key=lambda x: x['date'])
            if lpr_list:
                print(f"Tushare获取LPR成功: {len(lpr_list)}条记录")
                return lpr_list

    except Exception as e:
        print(f"Tushare获取LPR失败: {e}")

    return None


def load_lpr_from_cache():
    """从本地文件加载LPR数据"""
    for path in (get_cache_lpr_path(), LPR_DATA_FILE):
        if not os.path.exists(path):
            continue
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return None


def fetch_lpr_data(token=None):
    """优先使用中国货币网公开接口，失败后再尝试Tushare。"""
    global LPR_FETCH_ERROR
    errors = []

    try:
        data = fetch_chinamoney_lpr_data()
        LPR_FETCH_ERROR = ""
        print(f"中国货币网获取LPR成功: {len(data)}条记录")
        return data
    except Exception as e:
        errors.append(f"中国货币网: {e}")

    token = (token or get_tushare_token()).strip()
    if token:
        try:
            data = fetch_tushare_lpr_data(token)
            if data:
                LPR_FETCH_ERROR = ""
                return data
        except Exception as e:
            errors.append(f"Tushare: {e}")
    else:
        errors.append("Tushare: 未设置Token")

    LPR_FETCH_ERROR = "；".join(errors)
    print(f"LPR获取失败: {LPR_FETCH_ERROR}")
    return None


def save_lpr_to_cache(lpr_data):
    """保存LPR数据到本地文件"""
    try:
        with open(get_cache_lpr_path(), 'w', encoding='utf-8') as f:
            json.dump(lpr_data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存LPR数据失败: {e}")


def sort_lpr_data(lpr_data):
    if not lpr_data:
        return []
    return sorted(lpr_data, key=lambda item: item.get('date', ''))


class InterestCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("利息计算器")
        self.root.geometry("900x700")
        self.root.resizable(True, True)

        # LPR数据
        self.lpr_data = None
        self.load_or_fetch_lpr_data()

        # 样式
        self.setup_styles()

        # 变量
        self.principal_var = tk.StringVar()
        self.annual_rate_var = tk.StringVar()
        self.monthly_rate_var = tk.StringVar()
        self.use_lpr_var = tk.BooleanVar()
        self.lpr_multiplier_var = tk.StringVar(value="4")

        # 创建界面
        self.create_widgets()

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.configure('Title.TLabel', font=('Arial', 12, 'bold'))
        self.style.configure('Result.TLabel', font=('Arial', 11))
        self.style.configure('Input.TEntry', font=('Arial', 11))

    def load_or_fetch_lpr_data(self):
        """加载或获取LPR数据"""
        # 先尝试从本地加载缓存数据
        self.lpr_data = sort_lpr_data(load_lpr_from_cache())

        # 启动时不阻塞联网；如需最新数据，用户可点击“更新LPR”。
        if not self.lpr_data:
            self.lpr_data = BUILTIN_LPR_DATA.copy()
        self.lpr_data = sort_lpr_data(self.lpr_data)

    def update_lpr_data(self):
        """手动更新LPR数据"""
        online_data = fetch_lpr_data()
        if online_data:
            self.lpr_data = sort_lpr_data(online_data)
            save_lpr_to_cache(self.lpr_data)
            # 更新显示
            latest_lpr = self.lpr_data[-1]
            lpr_info = f"当前LPR(1年期): {latest_lpr.get('lpr', 3.45)}% ({latest_lpr.get('date', '')})"
            self.lpr_info_label.config(text=lpr_info)
            messagebox.showinfo("成功", "LPR数据已更新")
        else:
            # 获取失败，保留原有数据，显示警告
            if self.lpr_data:
                latest_lpr = self.lpr_data[-1]
                lpr_info = f"当前LPR(1年期): {latest_lpr.get('lpr', 3.45)}% ({latest_lpr.get('date', '')}) [缓存]"
                self.lpr_info_label.config(text=lpr_info)
            detail = f"\n\n失败原因：{LPR_FETCH_ERROR}" if LPR_FETCH_ERROR else ""
            messagebox.showwarning("警告", f"无法获取最新LPR数据，已继续使用本地缓存数据。{detail}")

    def prompt_tushare_token(self):
        """提示用户设置Tushare Token，避免打包版依赖环境变量。"""
        token = simpledialog.askstring(
            "设置 Tushare Token",
            "通常无需设置Token，程序会优先使用中国货币网公开数据。\n"
            "如果公开数据源暂时不可用，可在这里填写 Tushare Token 作为备用：",
            parent=self.root
        )
        if token and token.strip():
            token = token.strip()
            if save_tushare_token(token):
                return token
            messagebox.showerror("错误", "Token 保存失败，请检查用户目录写入权限。")
        return None

    def create_widgets(self):
        """创建所有界面组件"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 借款信息区域
        self.create_loan_info_frame(main_frame)

        # 利率设置区域
        self.create_rate_frame(main_frame)

        # 还款明细表区域
        self.create_repayment_frame(main_frame)

        # 按钮区域
        self.create_button_frame(main_frame)

        # 汇总输出区域
        self.create_summary_frame(main_frame)

    def create_loan_info_frame(self, parent):
        """借款信息区域"""
        frame = ttk.LabelFrame(parent, text="借款信息", padding="10")
        frame.pack(fill=tk.X, pady=(0, 10))

        row = ttk.Frame(frame)
        row.pack(fill=tk.X)

        ttk.Label(row, text="本金:").pack(side=tk.LEFT, padx=(0, 5))
        self.principal_entry = ttk.Entry(row, textvariable=self.principal_var, width=15, style='Input.TEntry')
        self.principal_entry.pack(side=tk.LEFT, padx=(0, 20))

        ttk.Label(row, text="起始日期:").pack(side=tk.LEFT, padx=(0, 5))
        self.start_date_entry = DateEntry(row, width=12, dateformat='%Y-%m-%d')
        self.start_date_entry.pack(side=tk.LEFT)

    def create_rate_frame(self, parent):
        """利率设置区域"""
        frame = ttk.LabelFrame(parent, text="利率设置", padding="10")
        frame.pack(fill=tk.X, pady=(0, 10))

        row1 = ttk.Frame(frame)
        row1.pack(fill=tk.X, pady=(0, 5))

        ttk.Label(row1, text="年利率(%):").pack(side=tk.LEFT, padx=(0, 5))
        self.annual_rate_entry = ttk.Entry(row1, textvariable=self.annual_rate_var, width=10, style='Input.TEntry')
        self.annual_rate_entry.pack(side=tk.LEFT, padx=(0, 20))

        ttk.Label(row1, text="月利率(%):").pack(side=tk.LEFT, padx=(0, 5))
        self.monthly_rate_entry = ttk.Entry(row1, textvariable=self.monthly_rate_var, width=10, style='Input.TEntry')
        self.monthly_rate_entry.pack(side=tk.LEFT)

        row2 = ttk.Frame(frame)
        row2.pack(fill=tk.X)

        self.use_lpr_check = ttk.Checkbutton(row2, text="使用LPR利率", variable=self.use_lpr_var)
        self.use_lpr_check.pack(side=tk.LEFT, padx=(0, 10))

        ttk.Label(row2, text="倍数:").pack(side=tk.LEFT, padx=(0, 5))
        self.lpr_multiplier_entry = ttk.Entry(row2, textvariable=self.lpr_multiplier_var, width=5, style='Input.TEntry')
        self.lpr_multiplier_entry.pack(side=tk.LEFT)

        # LPR信息显示
        if self.lpr_data:
            latest_lpr = self.lpr_data[-1] if self.lpr_data else {'lpr': 3.45, 'date': '2024-12-20'}
            lpr_info = f"当前LPR(1年期): {latest_lpr.get('lpr', 3.45)}% ({latest_lpr.get('date', '')})"
            self.lpr_info_label = ttk.Label(row2, text=lpr_info, foreground='blue')
            self.lpr_info_label.pack(side=tk.LEFT, padx=(20, 10))

        ttk.Button(row2, text="更新LPR", command=self.update_lpr_data).pack(side=tk.LEFT)
        ttk.Button(row2, text="设置Token", command=self.configure_tushare_token).pack(side=tk.LEFT, padx=(8, 0))

    def configure_tushare_token(self):
        """手动设置或更新Tushare Token。"""
        current_token = get_tushare_token()
        token = simpledialog.askstring(
            "设置 Tushare Token",
            "请输入 Tushare Token：",
            initialvalue=current_token,
            parent=self.root
        )
        if token is None:
            return
        token = token.strip()
        if not token:
            messagebox.showwarning("提示", "Token 不能为空。如暂不更新，可直接使用内置或缓存LPR数据。")
            return
        if save_tushare_token(token):
            messagebox.showinfo("成功", "Tushare Token 已保存。现在可以点击“更新LPR”。")
        else:
            messagebox.showerror("错误", "Token 保存失败，请检查用户目录写入权限。")

    def create_repayment_frame(self, parent):
        """还款明细表区域"""
        frame = ttk.LabelFrame(parent, text="资金流水明细", padding="10")
        frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 创建表格
        columns = ('序号', '日期', '类型', '金额', '应还利息', '实还利息', '当期欠息', '当期剩余本金')
        self.repayment_tree = ttk.Treeview(frame, columns=columns, show='headings', height=8)

        # 设置列宽
        self.repayment_tree.heading('序号', text='序号')
        self.repayment_tree.heading('日期', text='日期')
        self.repayment_tree.heading('类型', text='类型')
        self.repayment_tree.heading('金额', text='金额')
        self.repayment_tree.heading('应还利息', text='应还利息')
        self.repayment_tree.heading('实还利息', text='实还利息')
        self.repayment_tree.heading('当期欠息', text='当期欠息')
        self.repayment_tree.heading('当期剩余本金', text='当期剩余本金')

        self.repayment_tree.column('序号', width=45, anchor='center')
        self.repayment_tree.column('日期', width=110, anchor='center')
        self.repayment_tree.column('类型', width=50, anchor='center')
        self.repayment_tree.column('金额', width=100, anchor='e')
        self.repayment_tree.column('应还利息', width=100, anchor='e')
        self.repayment_tree.column('实还利息', width=100, anchor='e')
        self.repayment_tree.column('当期欠息', width=100, anchor='e')
        self.repayment_tree.column('当期剩余本金', width=120, anchor='e')

        # 滚动条
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=self.repayment_tree.yview)
        self.repayment_tree.configure(yscrollcommand=scrollbar.set)

        self.repayment_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 存储还款行数据
        self.repayment_rows = []

    def create_button_frame(self, parent):
        """按钮区域"""
        frame = ttk.Frame(parent)
        frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(frame, text="添加还款", command=self.add_repayment_row).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="添加借款", command=self.add_borrowing_row).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="修改", command=self.edit_repayment_row).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="删除", command=self.delete_repayment_row).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="计算过程", command=self.show_calculation_detail).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="清空", command=self.clear_repayments).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="导出Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(frame, text="汇总结果", command=self.calculate).pack(side=tk.RIGHT)

    def create_summary_frame(self, parent):
        """汇总输出区域"""
        frame = ttk.LabelFrame(parent, text="汇总输出", padding="10")
        frame.pack(fill=tk.X)

        row = ttk.Frame(frame)
        row.pack(fill=tk.X)

        ttk.Label(row, text="剩余本金:", font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.remaining_principal_label = ttk.Label(row, text="0.00", font=('Arial', 11), width=15)
        self.remaining_principal_label.pack(side=tk.LEFT, padx=(0, 30))

        ttk.Label(row, text="欠付利息:", font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.arrears_interest_label = ttk.Label(row, text="0.00", font=('Arial', 11), width=15)
        self.arrears_interest_label.pack(side=tk.LEFT, padx=(0, 30))

        ttk.Label(row, text="尚欠本息:", font=('Arial', 11, 'bold')).pack(side=tk.LEFT, padx=(0, 5))
        self.total_arrears_label = ttk.Label(row, text="0.00", font=('Arial', 11), width=15)
        self.total_arrears_label.pack(side=tk.LEFT)

    def add_repayment_row(self):
        """添加一行还款记录 - 弹出输入对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加还款记录")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中
        dialog.geometry("+{}+{}".format(
            self.root.winfo_x() + self.root.winfo_width() // 2 - 150,
            self.root.winfo_y() + self.root.winfo_height() // 2 - 75
        ))

        # 日期
        ttk.Label(dialog, text="还款日期:").grid(row=0, column=0, padx=10, pady=10, sticky='e')
        date_entry = DateEntry(dialog, width=12, dateformat='%Y-%m-%d')
        date_entry.grid(row=0, column=1, padx=10, pady=10)

        # 金额
        ttk.Label(dialog, text="还款金额:").grid(row=1, column=0, padx=10, pady=10, sticky='e')
        amount_var = tk.StringVar()
        amount_entry = ttk.Entry(dialog, textvariable=amount_var, width=15)
        amount_entry.grid(row=1, column=1, padx=10, pady=10)

        def on_confirm():
            try:
                repayment_date = date_entry.get_date()
                amount = to_decimal(amount_var.get())

                if repayment_date < self.start_date_entry.get_date():
                    messagebox.showerror("输入错误", "还款日期不能早于起始日期", parent=dialog)
                    return
                if amount <= 0:
                    messagebox.showerror("输入错误", "还款金额必须大于0", parent=dialog)
                    return

                # 存储行数据
                row_data = {
                    'num': len(self.repayment_rows) + 1,
                    'date': repayment_date,
                    'type': 'repay',
                    'amount': amount,
                    'interest_due': 0,
                    'interest_paid': 0,
                    'current_arrears': 0,
                    'remaining_principal': 0,
                    'calculation_detail': ''
                }
                self.repayment_rows.append(row_data)

                # 重新排序并更新序号
                self.repayment_rows.sort(key=lambda x: x['date'])
                for i, row in enumerate(self.repayment_rows):
                    row['num'] = i + 1

                # 重新计算所有行，避免插入较早日期时后续行仍是旧结果
                self.recalculate_all()
                dialog.destroy()
            except ValueError:
                messagebox.showerror("输入错误", "请输入有效的金额", parent=dialog)

        def on_cancel():
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="确定", command=on_confirm).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.LEFT, padx=5)

        amount_entry.focus()
        dialog.bind('<Return>', lambda e: on_confirm())
        dialog.bind('<Escape>', lambda e: on_cancel())

    def add_borrowing_row(self):
        """添加一笔借款记录 - 弹出输入对话框"""
        dialog = tk.Toplevel(self.root)
        dialog.title("添加借款记录")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中
        dialog.geometry("+{}+{}".format(
            self.root.winfo_x() + self.root.winfo_width() // 2 - 150,
            self.root.winfo_y() + self.root.winfo_height() // 2 - 75
        ))

        # 日期
        ttk.Label(dialog, text="借款日期:").grid(row=0, column=0, padx=10, pady=10, sticky='e')
        date_entry = DateEntry(dialog, width=12, dateformat='%Y-%m-%d')
        date_entry.grid(row=0, column=1, padx=10, pady=10)

        # 金额
        ttk.Label(dialog, text="借款金额:").grid(row=1, column=0, padx=10, pady=10, sticky='e')
        amount_var = tk.StringVar()
        amount_entry = ttk.Entry(dialog, textvariable=amount_var, width=15)
        amount_entry.grid(row=1, column=1, padx=10, pady=10)

        def on_confirm():
            try:
                borrow_date = date_entry.get_date()
                amount = to_decimal(amount_var.get())

                if borrow_date < self.start_date_entry.get_date():
                    messagebox.showerror("输入错误", "借款日期不能早于起始日期", parent=dialog)
                    return
                if amount <= 0:
                    messagebox.showerror("输入错误", "借款金额必须大于0", parent=dialog)
                    return

                row_data = {
                    'num': len(self.repayment_rows) + 1,
                    'date': borrow_date,
                    'type': 'borrow',
                    'amount': amount,
                    'interest_due': 0,
                    'interest_paid': 0,
                    'current_arrears': 0,
                    'remaining_principal': 0,
                    'calculation_detail': ''
                }
                self.repayment_rows.append(row_data)

                # 重新排序并更新序号
                self.repayment_rows.sort(key=lambda x: x['date'])
                for i, row in enumerate(self.repayment_rows):
                    row['num'] = i + 1

                # 重新计算所有行
                self.recalculate_all()
                dialog.destroy()
            except ValueError:
                messagebox.showerror("输入错误", "请输入有效的金额", parent=dialog)

        def on_cancel():
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="确定", command=on_confirm).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.LEFT, padx=5)

        amount_entry.focus()
        dialog.bind('<Return>', lambda e: on_confirm())
        dialog.bind('<Escape>', lambda e: on_cancel())

    def edit_repayment_row(self):
        """修改选中的记录（借款或还款）"""
        selection = self.repayment_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要修改的行")
            return

        # 获取选中的行
        item = selection[0]
        values = self.repayment_tree.item(item, 'values')
        row_num = int(values[0])

        # 找到对应的行数据
        row_data = None
        for row in self.repayment_rows:
            if row['num'] == row_num:
                row_data = row
                break

        if not row_data:
            messagebox.showerror("错误", "未找到对应的记录")
            return

        is_borrow = row_data.get('type') == 'borrow'
        type_label = "借款" if is_borrow else "还款"

        # 弹出修改对话框
        dialog = tk.Toplevel(self.root)
        dialog.title(f"修改{type_label}记录")
        dialog.geometry("300x150")
        dialog.transient(self.root)
        dialog.grab_set()

        # 居中
        dialog.geometry("+{}+{}".format(
            self.root.winfo_x() + self.root.winfo_width() // 2 - 150,
            self.root.winfo_y() + self.root.winfo_height() // 2 - 75
        ))

        # 日期
        ttk.Label(dialog, text=f"{type_label}日期:").grid(row=0, column=0, padx=10, pady=10, sticky='e')
        date_entry = DateEntry(dialog, width=12, dateformat='%Y-%m-%d')
        date_entry.set_date(row_data['date'])
        date_entry.grid(row=0, column=1, padx=10, pady=10)

        # 金额
        ttk.Label(dialog, text=f"{type_label}金额:").grid(row=1, column=0, padx=10, pady=10, sticky='e')
        amount_var = tk.StringVar(value=money_text(row_data['amount']))
        amount_entry = ttk.Entry(dialog, textvariable=amount_var, width=15)
        amount_entry.grid(row=1, column=1, padx=10, pady=10)

        def on_confirm():
            try:
                new_date = date_entry.get_date()
                amount = to_decimal(amount_var.get())

                if new_date < self.start_date_entry.get_date():
                    messagebox.showerror("输入错误", f"{type_label}日期不能早于起始日期", parent=dialog)
                    return
                if is_borrow and amount <= 0:
                    messagebox.showerror("输入错误", "借款金额必须大于0", parent=dialog)
                    return
                if not is_borrow and amount <= 0:
                    messagebox.showerror("输入错误", "还款金额必须大于0", parent=dialog)
                    return

                # 更新数据
                row_data['date'] = new_date
                row_data['amount'] = amount

                # 重新排序并更新序号
                self.repayment_rows.sort(key=lambda x: x['date'])
                for i, row in enumerate(self.repayment_rows):
                    row['num'] = i + 1

                # 重新计算所有行
                self.recalculate_all()
                dialog.destroy()
            except ValueError:
                messagebox.showerror("输入错误", "请输入有效的金额", parent=dialog)

        def on_cancel():
            dialog.destroy()

        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(btn_frame, text="确定", command=on_confirm).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="取消", command=on_cancel).pack(side=tk.LEFT, padx=5)

        amount_entry.focus()
        amount_entry.select_range(0, tk.END)
        dialog.bind('<Return>', lambda e: on_confirm())
        dialog.bind('<Escape>', lambda e: on_cancel())

    def delete_repayment_row(self):
        """删除选中的记录"""
        selection = self.repayment_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要删除的行")
            return

        item = selection[0]
        values = self.repayment_tree.item(item, 'values')
        row_num = int(values[0])

        if messagebox.askyesno("确认", f"确定要删除第 {row_num} 条记录吗?"):
            # 找到并删除
            for i, row in enumerate(self.repayment_rows):
                if row['num'] == row_num:
                    self.repayment_rows.pop(i)
                    break

            # 重新排序并更新序号
            self.repayment_rows.sort(key=lambda x: x['date'])
            for i, row in enumerate(self.repayment_rows):
                row['num'] = i + 1

            # 重新计算所有行
            self.recalculate_all()

    def recalculate_all(self):
        """重新计算所有行的数据"""
        if not self.repayment_rows:
            self.refresh_treeview()
            return

        try:
            principal = to_decimal(self.principal_var.get())
        except ValueError:
            principal = Decimal("0")

        start_date = self.start_date_entry.get_date()
        try:
            annual_rate = to_decimal(self.annual_rate_var.get(), Decimal("0")) / Decimal("100")
            monthly_rate = to_decimal(self.monthly_rate_var.get(), Decimal("0")) / Decimal("100")
            multiplier = to_decimal(self.lpr_multiplier_var.get(), Decimal("4"))
        except ValueError:
            annual_rate = Decimal("0")
            monthly_rate = Decimal("0")
            multiplier = Decimal("4")
        user_rate = max(annual_rate, monthly_rate * 12)
        use_lpr = self.use_lpr_var.get()

        prev_date = start_date
        prev_remaining = principal
        brought_forward_arrears = Decimal("0")

        for row in self.repayment_rows:
            # 计算本期应还利息（按当前剩余本金、上期日期到本期日期）
            interest_due, segments = self.calculate_segment_interest(
                prev_date, row['date'], prev_remaining, user_rate, use_lpr, multiplier, include_details=True
            )
            total_interest_due = interest_due + brought_forward_arrears

            if row.get('type') == 'borrow':
                # 借款：利息全部挂账，本金增加
                row['interest_due'] = interest_due
                row['interest_paid'] = Decimal("0")
                row['current_arrears'] = total_interest_due
                row['remaining_principal'] = prev_remaining + row['amount']
                row['calculation_detail'] = self.build_calculation_detail(
                    row, prev_date, prev_remaining, brought_forward_arrears, segments,
                    total_interest_due, Decimal("0"), Decimal("0"), row['remaining_principal']
                )
            else:
                # 还款：先息后本
                interest_paid = min(row['amount'], total_interest_due)
                principal_paid = min(row['amount'] - interest_paid, max(prev_remaining, Decimal("0")))
                current_arrears = total_interest_due - interest_paid
                remaining_principal = prev_remaining - principal_paid

                row['interest_due'] = interest_due
                row['interest_paid'] = interest_paid
                row['current_arrears'] = current_arrears
                row['remaining_principal'] = remaining_principal
                row['calculation_detail'] = self.build_calculation_detail(
                    row, prev_date, prev_remaining, brought_forward_arrears, segments,
                    total_interest_due, interest_paid, principal_paid, remaining_principal
                )

            prev_date = row['date']
            prev_remaining = row['remaining_principal']
            brought_forward_arrears = row['current_arrears']

        self.refresh_treeview()
        self.update_summary()

    def refresh_treeview(self):
        """刷新treeview显示"""
        # 清空所有项
        for item in self.repayment_tree.get_children():
            self.repayment_tree.delete(item)

        # 重新插入所有行
        for row in self.repayment_rows:
            type_text = "借款" if row.get('type') == 'borrow' else "还款"
            values = (
                row['num'],
                row['date'].strftime('%Y-%m-%d') if row['date'] else '',
                type_text,
                money_text(row['amount']) if row['amount'] else '',
                money_text(row['interest_due']),
                money_text(row['interest_paid']),
                money_text(row['current_arrears']),
                money_text(row['remaining_principal'])
            )
            item = self.repayment_tree.insert('', 'end', values=values)
            # 借款行用蓝色标记
            if row.get('type') == 'borrow':
                self.repayment_tree.tag_configure('borrow', foreground='blue')
                self.repayment_tree.item(item, tags=('borrow',))

        # 更新汇总显示
        self.update_summary()

    def update_summary(self):
        """更新汇总显示"""
        if not self.repayment_rows:
            self.remaining_principal_label.config(text="0.00")
            self.arrears_interest_label.config(text="0.00")
            self.total_arrears_label.config(text="0.00")
            return

        last_row = self.repayment_rows[-1]
        remaining_principal = last_row['remaining_principal']
        arrears_interest = last_row['current_arrears']
        total_arrears = remaining_principal + arrears_interest

        self.remaining_principal_label.config(text=money_text(remaining_principal))
        self.arrears_interest_label.config(text=money_text(arrears_interest))
        self.total_arrears_label.config(text=money_text(total_arrears))

    def build_calculation_detail(self, row, prev_date, prev_principal, brought_forward_arrears,
                                 segments, total_interest_due, interest_paid,
                                 principal_paid, remaining_principal):
        """生成供人工复验的单条流水计算过程。"""
        type_text = "借款" if row.get('type') == 'borrow' else "还款"
        lines = [
            f"第{row['num']}条 {type_text}记录",
            f"流水日期: {row['date'].strftime('%Y-%m-%d')}",
            f"流水金额: {money_text(row['amount'])}",
            f"上一计息日: {prev_date.strftime('%Y-%m-%d')}",
            f"起算本金: {money_text(prev_principal)}",
            f"上期结转欠息: {money_text(brought_forward_arrears)}",
            "",
            "本期利息计算:",
        ]

        if not segments:
            lines.append("无计息天数，本期新增利息 0.00。")
        else:
            for segment in segments:
                formula = (
                    f"{money_text(segment['principal'])} × {percent_text(segment['rate'])}% "
                    f"÷ 360 × {segment['days']}天 = {money_text(segment['interest'])}"
                )
                lines.extend([
                    f"{segment['start'].strftime('%Y-%m-%d')} 至 {segment['end'].strftime('%Y-%m-%d')}:",
                    f"  适用上限: {segment['cap_description']}",
                    f"  计算公式: {formula}",
                ])

        current_interest = total_interest_due - brought_forward_arrears
        lines.extend([
            "",
            f"本期新增应还利息: {money_text(current_interest)}",
            f"本期应处理利息合计: {money_text(current_interest)} + {money_text(brought_forward_arrears)} = {money_text(total_interest_due)}",
        ])

        if row.get('type') == 'borrow':
            lines.extend([
                "本条为追加借款，不冲抵利息和本金。",
                f"剩余本金: {money_text(prev_principal)} + {money_text(row['amount'])} = {money_text(remaining_principal)}",
                f"结转欠息: {money_text(total_interest_due)}",
            ])
        else:
            excess_paid = row['amount'] - interest_paid - principal_paid
            lines.extend([
                "本条为还款，按先息后本处理。",
                f"实还利息: min({money_text(row['amount'])}, {money_text(total_interest_due)}) = {money_text(interest_paid)}",
                f"冲抵本金: min({money_text(row['amount'])} - {money_text(interest_paid)}, {money_text(prev_principal)}) = {money_text(principal_paid)}",
                f"剩余本金: {money_text(prev_principal)} - {money_text(principal_paid)} = {money_text(remaining_principal)}",
                f"结转欠息: {money_text(total_interest_due)} - {money_text(interest_paid)} = {money_text(row['current_arrears'])}",
            ])
            if excess_paid > 0:
                lines.append(f"超出本息金额: {money_text(excess_paid)}（未继续冲抵为负本金）")

        return "\n".join(lines)

    def show_calculation_detail(self):
        """显示选中流水的计算过程。"""
        selection = self.repayment_tree.selection()
        if not selection:
            messagebox.showwarning("警告", "请先选择要查看的行")
            return

        item = selection[0]
        values = self.repayment_tree.item(item, 'values')
        row_num = int(values[0])
        row_data = next((row for row in self.repayment_rows if row['num'] == row_num), None)
        if not row_data:
            messagebox.showerror("错误", "未找到对应的记录")
            return

        if not row_data.get('calculation_detail'):
            self.recalculate_all()
            row_data = next((row for row in self.repayment_rows if row['num'] == row_num), row_data)

        dialog = tk.Toplevel(self.root)
        dialog.title(f"第{row_num}条计算过程")
        dialog.geometry("720x460")
        dialog.transient(self.root)

        text = tk.Text(dialog, wrap=tk.WORD, font=('Consolas', 10))
        scrollbar = ttk.Scrollbar(dialog, orient=tk.VERTICAL, command=text.yview)
        text.configure(yscrollcommand=scrollbar.set)
        text.insert("1.0", row_data.get('calculation_detail', '暂无计算过程，请先汇总结果。'))
        text.configure(state=tk.DISABLED)
        text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 10), pady=10)

    def export_to_excel(self):
        """导出计算结果到Excel"""
        if not self.repayment_rows:
            messagebox.showwarning("警告", "没有可导出的记录")
            return

        # 获取保存路径
        file_path = filedialog.asksaveasfilename(
            defaultextension='.xlsx',
            filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')],
            title='保存Excel文件'
        )
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "利息计算结果"

            # 样式定义
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')

            # 借款信息区域
            ws['A1'] = '借款本金'
            ws['B1'] = decimal_to_excel_number(self.principal_var.get()) if self.principal_var.get() else 0
            ws['C1'] = '起始日期'
            ws['D1'] = self.start_date_entry.get_date().strftime('%Y-%m-%d') if self.start_date_entry.get_date() else ''

            ws['A2'] = '年利率(%)'
            ws['B2'] = float(to_decimal(self.annual_rate_var.get(), Decimal("0")))
            ws['C2'] = '月利率(%)'
            ws['D2'] = float(to_decimal(self.monthly_rate_var.get(), Decimal("0")))
            ws['E2'] = '使用LPR'
            ws['F2'] = '是' if self.use_lpr_var.get() else '否'
            ws['G2'] = '倍数'
            ws['H2'] = float(to_decimal(self.lpr_multiplier_var.get(), Decimal("4")))

            # 明细表标题行
            headers = ['序号', '日期', '类型', '金额', '应还利息', '实还利息', '当期欠息', '当期剩余本金', '计算过程']
            header_row = 4
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = center_align

            # 数据行
            for row_idx, row_data in enumerate(self.repayment_rows, start=header_row + 1):
                type_text = "借款" if row_data.get('type') == 'borrow' else "还款"

                ws.cell(row=row_idx, column=1, value=row_data['num']).border = thin_border
                ws.cell(row=row_idx, column=1).alignment = center_align
                ws.cell(row=row_idx, column=2, value=row_data['date'].strftime('%Y-%m-%d')).border = thin_border
                ws.cell(row=row_idx, column=2).alignment = center_align
                ws.cell(row=row_idx, column=3, value=type_text).border = thin_border
                ws.cell(row=row_idx, column=3).alignment = center_align
                ws.cell(row=row_idx, column=4, value=decimal_to_excel_number(row_data['amount'])).border = thin_border
                ws.cell(row=row_idx, column=4).alignment = right_align
                ws.cell(row=row_idx, column=4).number_format = '0.00'
                ws.cell(row=row_idx, column=5, value=decimal_to_excel_number(row_data['interest_due'])).border = thin_border
                ws.cell(row=row_idx, column=5).alignment = right_align
                ws.cell(row=row_idx, column=5).number_format = '0.00'
                ws.cell(row=row_idx, column=6, value=decimal_to_excel_number(row_data['interest_paid'])).border = thin_border
                ws.cell(row=row_idx, column=6).alignment = right_align
                ws.cell(row=row_idx, column=6).number_format = '0.00'
                ws.cell(row=row_idx, column=7, value=decimal_to_excel_number(row_data['current_arrears'])).border = thin_border
                ws.cell(row=row_idx, column=7).alignment = right_align
                ws.cell(row=row_idx, column=7).number_format = '0.00'
                ws.cell(row=row_idx, column=8, value=decimal_to_excel_number(row_data['remaining_principal'])).border = thin_border
                ws.cell(row=row_idx, column=8).alignment = right_align
                ws.cell(row=row_idx, column=8).number_format = '0.00'
                ws.cell(row=row_idx, column=9, value=row_data.get('calculation_detail', '')).border = thin_border
                ws.cell(row=row_idx, column=9).alignment = Alignment(wrap_text=True, vertical='top')

            # 汇总行
            summary_row = header_row + len(self.repayment_rows) + 1
            ws.cell(row=summary_row, column=1, value='汇总').font = header_font
            ws.cell(row=summary_row, column=4, value=decimal_to_excel_number(self.remaining_principal_label.cget("text"))).font = header_font
            ws.cell(row=summary_row, column=4).number_format = '0.00'
            ws.cell(row=summary_row, column=7, value=decimal_to_excel_number(self.arrears_interest_label.cget("text"))).font = header_font
            ws.cell(row=summary_row, column=7).number_format = '0.00'
            ws.cell(row=summary_row, column=8, value=decimal_to_excel_number(self.total_arrears_label.cget("text"))).font = header_font
            ws.cell(row=summary_row, column=8).number_format = '0.00'

            # 设置列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 8
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 70

            wb.save(file_path)
            messagebox.showinfo("成功", f"已导出到:\n{file_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败:\n{str(e)}")

    def clear_repayments(self):
        """清空所有记录"""
        for item in self.repayment_tree.get_children():
            self.repayment_tree.delete(item)
        self.repayment_rows = []
        self.update_summary()

    def validate_inputs(self):
        """验证输入"""
        try:
            principal = to_decimal(self.principal_var.get())
            if principal <= 0:
                messagebox.showerror("输入错误", "本金必须大于0")
                return False
        except ValueError:
            messagebox.showerror("输入错误", "请输入有效的本金金额")
            return False

        start_date = self.start_date_entry.get_date()
        if not start_date:
            messagebox.showerror("输入错误", "请选择起始日期")
            return False

        try:
            annual_rate = to_decimal(self.annual_rate_var.get(), Decimal("0")) / Decimal("100")
            monthly_rate = to_decimal(self.monthly_rate_var.get(), Decimal("0")) / Decimal("100")
            user_rate = max(annual_rate, monthly_rate * 12)
            if user_rate <= 0:
                messagebox.showerror("输入错误", "请输入有效的利率")
                return False
        except ValueError:
            messagebox.showerror("输入错误", "请输入有效的利率值")
            return False

        if self.use_lpr_var.get():
            try:
                multiplier = to_decimal(self.lpr_multiplier_var.get())
                if multiplier <= 0:
                    messagebox.showerror("输入错误", "倍数必须大于0")
                    return False
            except ValueError:
                messagebox.showerror("输入错误", "请输入有效的倍数")
                return False

        # 检查事件记录
        valid_events = []
        for row in self.repayment_rows:
            if row['date'] and row['amount'] > 0:
                if row['date'] < start_date:
                    messagebox.showerror("输入错误", "流水日期不能早于起始日期")
                    return False
                valid_events.append(row)

        if not valid_events:
            messagebox.showerror("输入错误", "请至少添加一笔借款或还款记录")
            return False

        # 按日期排序
        valid_events.sort(key=lambda x: x['date'])

        return True

    def calculate_segment_interest(self, start_date, end_date, principal, user_rate, use_lpr, multiplier, include_details=False):
        """
        计算一段时间的利息（考虑分段）
        """
        # 收集所有分割点
        split_points = set([start_date, end_date])

        # 添加法定利率规则分割点，按半开区间计算。
        for legal_point in (DATE_START_LEGAL, DATE_START_LPR):
            if start_date < legal_point < end_date:
                split_points.add(legal_point)

        # 添加LPR变动点
        if self.lpr_data and (use_lpr or end_date > DATE_START_LPR):
            for record in self.lpr_data:
                record_date = datetime.strptime(record['date'], '%Y-%m-%d').date()
                if start_date < record_date < end_date:
                    split_points.add(record_date)

        # 排序
        split_points = sorted(list(split_points))

        total_interest = Decimal("0")
        remaining = to_decimal(principal, Decimal("0"))
        detail_segments = []

        for i in range(len(split_points) - 1):
            segment_start = split_points[i]
            segment_end = split_points[i + 1]
            days = days_between(segment_start, segment_end)

            if days <= 0:
                continue

            # 获取该段适用利率
            rate_info = get_rate_info_for_period(
                segment_start, segment_end, user_rate, use_lpr, multiplier, self.lpr_data
            )
            rate = rate_info["rate"]

            # 计算利息
            segment_interest = remaining * (rate / YEAR_DAYS) * Decimal(days)
            total_interest += segment_interest
            detail_segments.append({
                "start": segment_start,
                "end": segment_end,
                "days": days,
                "principal": remaining,
                "rate": rate,
                "cap_description": rate_info["cap_description"],
                "interest": segment_interest,
            })

        if include_details:
            return total_interest, detail_segments
        return total_interest

    def calculate(self):
        """执行计算"""
        if not self.validate_inputs():
            return

        principal = to_decimal(self.principal_var.get())
        start_date = self.start_date_entry.get_date()
        annual_rate = to_decimal(self.annual_rate_var.get(), Decimal("0")) / Decimal("100")
        monthly_rate = to_decimal(self.monthly_rate_var.get(), Decimal("0")) / Decimal("100")
        user_rate = max(annual_rate, monthly_rate * 12)
        use_lpr = self.use_lpr_var.get()
        multiplier = to_decimal(self.lpr_multiplier_var.get(), Decimal("4"))

        # 获取有效事件记录并排序
        valid_events = []
        for row in self.repayment_rows:
            if row['date'] and row['amount'] > 0:
                valid_events.append(row)
        valid_events.sort(key=lambda x: x['date'])

        # 计算每笔事件
        current_principal = principal
        current_date = start_date
        brought_forward_arrears = Decimal("0")

        for row in valid_events:
            event_date = row['date']
            event_amount = row['amount']

            # 计算应还利息（从上次日期到本次日期，基于当前剩余本金）
            interest_due, segments = self.calculate_segment_interest(
                current_date, event_date, current_principal, user_rate, use_lpr, multiplier, include_details=True
            )
            total_interest_due = interest_due + brought_forward_arrears

            if row.get('type') == 'borrow':
                # 借款：利息全部挂账，本金增加
                row['interest_due'] = interest_due
                row['interest_paid'] = Decimal("0")
                row['current_arrears'] = total_interest_due
                row['remaining_principal'] = current_principal + event_amount
                row['calculation_detail'] = self.build_calculation_detail(
                    row, current_date, current_principal, brought_forward_arrears, segments,
                    total_interest_due, Decimal("0"), Decimal("0"), row['remaining_principal']
                )
            else:
                # 还款：先息后本
                interest_paid = min(event_amount, total_interest_due)
                principal_paid = min(event_amount - interest_paid, max(current_principal, Decimal("0")))
                current_arrears = total_interest_due - interest_paid
                remaining_principal = current_principal - principal_paid

                row['interest_due'] = interest_due
                row['interest_paid'] = interest_paid
                row['current_arrears'] = current_arrears
                row['remaining_principal'] = remaining_principal
                row['calculation_detail'] = self.build_calculation_detail(
                    row, current_date, current_principal, brought_forward_arrears, segments,
                    total_interest_due, interest_paid, principal_paid, remaining_principal
                )

            current_date = event_date
            current_principal = row['remaining_principal']
            brought_forward_arrears = row['current_arrears']

        # 更新treeview显示（包括汇总）
        self.refresh_treeview()


def main():
    root = tk.Tk()
    app = InterestCalculatorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
